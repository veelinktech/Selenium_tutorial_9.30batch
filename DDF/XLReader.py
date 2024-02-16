import openpyxl


def total_rows(excelPath, sheetName):
    wb = openpyxl.load_workbook(excelPath)
    sheet = wb[sheetName]
    return sheet.max_row


def total_columns(excelPath, sheetName):
    wb = openpyxl.load_workbook(excelPath)
    sheet = wb[sheetName]
    return sheet.max_column


def read_Data(excelPath, sheetName, row_Num, col_Num):
    wb = openpyxl.load_workbook(excelPath)
    sheet = wb[sheetName]
    data = sheet.cell(row=row_Num, column=col_Num).value
    return data


def write_Data(excelPath, sheetName, row_Num, col_Num, val):
    wb = openpyxl.load_workbook(excelPath)
    sheet = wb[sheetName]
    sheet.cell(row=row_Num, column=col_Num).value = val
    wb.save(excelPath)
