import openpyxl

#Module name --> openpyxl
#Methdds --> load_workbook, cell
#Property --> active, max_row, max_column, value

wb = openpyxl.load_workbook('C:\\Users\\pc\\Desktop\\DDF.xlsx')
sheet = wb['Sheet1']
#sheet = wb.active
total_row = sheet.max_row
total_column = sheet.max_column

print("Row count = ", total_row)
print("Column count = ", total_column)

#single column
data = sheet.cell(row=5,column=1).value
print(data)

for r in range(1,total_row+1):
    for c in range(1,total_column+1):
        print(sheet.cell(row=r,column=c).value)